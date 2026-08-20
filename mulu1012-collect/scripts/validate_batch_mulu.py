#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""批次总目字段复检。

检查生成结果是否遵守字段指南：主表题名/别名分流、ID/行数、必填项、入口 URL、
受控字段和 AI 留痕。报告为 JSON，可作为上传前的阻断门。
"""
import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


MAIN_HEADERS = [
    '条目ID', '采集批次', '来源名称', '来源URL', '复核状态', '自建/特色库？', '史学数据库？',
    'WPID', '线上状态', '简体中文名', '网址类型', '载体', '主办方', '作者/编纂者', '母库',
    '系列', '基金项目', '上线时间', '入口页url', '介绍页url', '访问方式', '网站语言',
    '网站国家/地区', '授权机构（本批次）', '该馆访问方式', '机构内名称', '授权页url', '核验日期',
    '简介', '正式介绍', '官网介绍', '官网介绍出处url', '数据库类型', '电子资源形态', '主分类',
    '特色史料', '研究主题', '专门史', '时期', '国别/区域史', '材料类型', '材料语言', '收录年限',
    '评级', 'AI评分', 'AI评语', '备注'
]
URL_RE = re.compile(r'^https?://', re.I)
BAD_ENTRY_RE = re.compile(r'(?:ezproxy|(?:^|[./_-])sso(?:[./?_-]|$)|/han/|(?:^|/)login(?:[_./?-]|$))', re.I)
CJK_RE = re.compile(r'[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]')
JAPANESE_KANA_RE = re.compile(r'[\u3040-\u30FF\u31F0-\u31FF]')
HANGUL_RE = re.compile(r'[\u1100-\u11FF\u3130-\u318F\uAC00-\uD7AF]')
TRADITIONAL_ONLY_RE = re.compile(r'[臺灣學術體書報數據庫錄經傳漢語辭典總匯門類線網頁雜誌華東亞區會館標題彙與為後業務實際現時從來對於專門歷廣邊國圖館檔資料數位歷萬億啟開關聯繫統網絡叢專業舊龍馬藝術聲讀寫譯齊劉張陳吳趙雲廣門風義樂禮詩畫醫藥農產縣鄉鎮區劃戰爭黨團體機構權檢閱覽藏書目錄論文期刊資料庫電寶資訊]')
SIMPLIFIED_ONLY_RE = re.compile(r'[台湾学术体书报数据库录经传汉语辞典总汇门类线网页杂志华东亚区会馆标题汇与为后业务实际现时从来对于专门历广边国图馆档资料数位启开关联系统网络丛专业旧龙马艺术声读写译齐刘张陈吴赵云门风义乐礼诗画医药农产县乡镇区划战争党团机构权检阅览藏书目录论文期刊库]')
ALIAS_TYPES = {'原名', '译名', '文字变体', '缩写/代号', '描述性名称', '完整正题名', '副刊名', '曾用名', '其他'}
ALIAS_LANGUAGES = {'', '简体中文', '繁体中文', '英文', '日文', '韩文', '德文', '法文', '俄文', '越南文', '多语种', '其他'}
ACCESS_VALUES = {
    '图书馆访问', '图书馆访问＞仅馆内访问', '图书馆访问＞可馆外访问', '图书馆访问＞国外图书馆访问',
    '需授权', '需授权＞个人订购', '需授权＞机构订购', '需授权＞申请制', '需授权＞会员制',
    '需授权＞按次付费', '需授权＞内部资源', '免费', '无法访问', '待整理'
}


def nonempty(value):
    return value not in (None, '')


def text(value):
    return '' if value is None else str(value).strip()


def is_likely_simplified_title(value):
    value = text(value)
    return bool(
        value
        and CJK_RE.search(value)
        and not JAPANESE_KANA_RE.search(value)
        and not HANGUL_RE.search(value)
        and not TRADITIONAL_ONLY_RE.search(value)
        and SIMPLIFIED_ONLY_RE.search(value)
    )


def split_multiple(value):
    return [item.strip() for item in re.split(r'[、;；]', text(value)) if item.strip()]


def normalized_term(value):
    return re.sub(r'^[\s　]*└\s*', '', text(value))


def load_source(path):
    if not path:
        return []
    data = json.loads(Path(path).read_text(encoding='utf-8'))
    if not isinstance(data, list):
        raise ValueError('source JSON 顶层不是数组')
    return data


def main():
    parser = argparse.ArgumentParser(description='复检批次总目字段是否符合字段指南')
    parser.add_argument('--xlsx', required=True)
    parser.add_argument('--source-json')
    parser.add_argument('--title-translations')
    parser.add_argument('--report', required=True)
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    errors, warnings = [], []
    controlled_terms = {}
    if '词表速查' in wb.sheetnames:
        quick_rows = list(wb['词表速查'].iter_rows(values_only=True))
        if quick_rows:
            for col, header in enumerate(quick_rows[0]):
                if not nonempty(header):
                    continue
                controlled_terms[text(header)] = {
                    normalized_term(row[col]) for row in quick_rows[1:]
                    if col < len(row) and nonempty(row[col])
                }
    if '条目表' not in wb.sheetnames:
        errors.append({'code': 'missing_sheet', 'message': '缺少「条目表」'})
        rows = []
    else:
        ws = wb['条目表']
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0]) if rows else []
        if headers != MAIN_HEADERS:
            errors.append({'code': 'header_mismatch', 'message': '条目表表头不是 47 列目标顺序', 'actual': headers})
        index = {h: i for i, h in enumerate(headers)}
        data_rows = [row for row in rows[1:] if row and nonempty(row[0])]

        ids = [text(row[index.get('条目ID', 0)]) for row in data_rows]
        dup = sorted({item for item in ids if ids.count(item) > 1})
        if dup:
            errors.append({'code': 'duplicate_ids', 'count': len(dup), 'ids': dup[:20]})

        for required in ('网址类型', '载体', '入口页url', '访问方式', '简介', '主分类', 'AI评分', 'AI评语', '备注'):
            missing = [row_number for row_number, row in enumerate(data_rows, start=2) if not nonempty(row[index[required]])]
            if missing:
                errors.append({'code': 'missing_field', 'field': required, 'count': len(missing), 'rows': missing[:20]})

        for row_number, row in enumerate(data_rows, start=2):
            title = text(row[index['简体中文名']])
            if title and not is_likely_simplified_title(title):
                errors.append({'code': 'invalid_simplified_title', 'row': row_number, 'title': title})
            entry = text(row[index['入口页url']])
            if entry and not URL_RE.match(entry):
                errors.append({'code': 'bad_entry_url', 'row': row_number, 'value': entry})
            if entry and BAD_ENTRY_RE.search(entry):
                errors.append({'code': 'proxy_or_login_entry', 'row': row_number, 'value': entry})
            score = row[index['AI评分']]
            if not isinstance(score, (int, float)) or not 0 <= score <= 100:
                errors.append({'code': 'bad_ai_score', 'row': row_number, 'value': score})

        for field in ('简介', '正式介绍'):
            invalid_text = []
            for row_number, row in enumerate(data_rows, start=2):
                field_text = text(row[index[field]])
                if field_text and not is_likely_simplified_title(field_text):
                    invalid_text.append({'row': row_number, 'sample': field_text[:120]})
            if invalid_text:
                errors.append({
                    'code': 'non_simplified_chinese_text',
                    'field': field,
                    'count': len(invalid_text),
                    'sample': invalid_text[:10],
                })

        for field in ('网址类型', '载体', '数据库类型', '电子资源形态', '材料语言', '评级', '网站语言'):
            term_column = '别名语言' if field == '网站语言' else field
            allowed = set(controlled_terms.get(term_column, set()))
            if field == '网址类型':
                allowed -= {'网站', '微信资源'}
            if not allowed:
                continue
            invalid_values = {}
            for row in data_rows:
                for item in split_multiple(row[index[field]]):
                    if item not in allowed:
                        invalid_values[item] = invalid_values.get(item, 0) + 1
            if invalid_values:
                errors.append({
                    'code': 'invalid_controlled_value',
                    'field': field,
                    'count': sum(invalid_values.values()),
                    'values': dict(sorted(invalid_values.items(), key=lambda item: (-item[1], item[0]))[:30]),
                })

        invalid_access = {}
        for row in data_rows:
            for item in split_multiple(row[index['访问方式']]):
                if item not in ACCESS_VALUES:
                    invalid_access[item] = invalid_access.get(item, 0) + 1
        if invalid_access:
            errors.append({
                'code': 'invalid_controlled_value',
                'field': '访问方式',
                'count': sum(invalid_access.values()),
                'values': invalid_access,
            })

        if len(data_rows) == 0:
            errors.append({'code': 'empty_main_table'})

    alias_rows = []
    if '别名明细' not in wb.sheetnames:
        errors.append({'code': 'missing_sheet', 'message': '缺少「别名明细」'})
    else:
        alias_ws = wb['别名明细']
        alias_rows = list(alias_ws.iter_rows(values_only=True))
        alias_header = list(alias_rows[0]) if alias_rows else []
        expected_alias = ['条目ID', '名称', '类型', '语言']
        if alias_header[:4] != expected_alias:
            errors.append({'code': 'alias_header_mismatch', 'actual': alias_header})
        else:
            known_ids = set(ids) if 'ids' in locals() else set()
            for row_number, row in enumerate(alias_rows[1:], start=2):
                if not row or not nonempty(row[0]):
                    continue
                alias_id, alias_name, alias_type, alias_language = (text(row[i]) for i in range(4))
                if alias_id not in known_ids:
                    errors.append({'code': 'orphan_alias', 'row': row_number, 'id': alias_id})
                if not alias_name:
                    errors.append({'code': 'missing_alias_name', 'row': row_number, 'id': alias_id})
                if alias_type not in ALIAS_TYPES:
                    errors.append({'code': 'invalid_alias_type', 'row': row_number, 'id': alias_id, 'value': alias_type})
                if alias_language not in ALIAS_LANGUAGES:
                    errors.append({'code': 'invalid_alias_language', 'row': row_number, 'id': alias_id, 'value': alias_language})

    source = load_source(args.source_json)
    if source and len(source) != len(rows) - 1:
        errors.append({'code': 'source_count_mismatch', 'source': len(source), 'xlsx': len(rows) - 1})

    title_index = {text(row[0]): text(row[9]) for row in rows[1:] if row and nonempty(row[0])}
    aliases_by_id = {}
    alias_type_counts = {}
    alias_language_blank = 0
    for row in alias_rows[1:]:
        if not row or not nonempty(row[0]):
            continue
        alias_id, alias_name, alias_type, alias_language = (text(row[i]) for i in range(4))
        aliases_by_id.setdefault(alias_id, set()).add(alias_name)
        alias_type_counts[alias_type] = alias_type_counts.get(alias_type, 0) + 1
        if not alias_language:
            alias_language_blank += 1
    alias_ids = {text(row[0]) for row in alias_rows[1:] if row and nonempty(row[0])}
    no_title_ids = sorted([rid for rid, title in title_index.items() if not title])
    no_alias_ids = sorted([rid for rid in no_title_ids if rid not in alias_ids])
    if no_alias_ids:
        errors.append({'code': 'no_title_no_alias', 'count': len(no_alias_ids), 'ids': no_alias_ids[:20]})
    if no_title_ids:
        errors.append({
            'code': 'manual_translation_required',
            'field': '简体中文名',
            'count': len(no_title_ids),
            'ids': no_title_ids[:20],
            'message': '缺少可确认的简体中文名；原名已入别名，完成译写前禁止上传',
        })
        note_index = index.get('AI评语') if 'index' in locals() else None
        no_translation_note = []
        if note_index is not None:
            for row in rows[1:]:
                if not row or not nonempty(row[0]) or text(row[9]):
                    continue
                note = text(row[note_index])
                if '简体中文名' not in note and '人工译写' not in note:
                    no_translation_note.append(text(row[0]))
        if no_translation_note:
            errors.append({
                'code': 'missing_manual_translation_note',
                'count': len(no_translation_note),
                'ids': no_translation_note[:20],
            })

    untraced_source_titles = []
    for offset, record in enumerate(source, start=1):
        record_id = f'XA-{offset:03d}'
        source_title = text(record.get('title')) if isinstance(record, dict) else ''
        if source_title and source_title != title_index.get(record_id, '') and source_title not in aliases_by_id.get(record_id, set()):
            untraced_source_titles.append({'id': record_id, 'title': source_title})
    if untraced_source_titles:
        errors.append({
            'code': 'untraced_source_title',
            'count': len(untraced_source_titles),
            'sample': untraced_source_titles[:20],
        })

    translated_titles = []
    if args.title_translations:
        translation_data = json.loads(Path(args.title_translations).read_text(encoding='utf-8'))
        if isinstance(translation_data, list):
            translated_titles = translation_data
        elif isinstance(translation_data, dict):
            translated_titles = [dict({'id': key}, **value) if isinstance(value, dict) else {'id': key, 'zh': value} for key, value in translation_data.items()]
        else:
            errors.append({'code': 'bad_title_translation_file', 'message': '题名译写文件须为数组或对象'})
        translation_errors = []
        notes_by_id = {
            text(row[0]): text(row[index['AI评语']]) for row in rows[1:]
            if row and nonempty(row[0])
        }
        for item in translated_titles:
            record_id = text(item.get('id'))
            translated = text(item.get('zh'))
            if not record_id or not translated:
                translation_errors.append({'id': record_id, 'reason': 'missing_id_or_zh'})
                continue
            if title_index.get(record_id) != translated:
                translation_errors.append({'id': record_id, 'reason': 'workbook_title_mismatch', 'expected': translated, 'actual': title_index.get(record_id)})
            if 'AI' not in notes_by_id.get(record_id, '') or '译写' not in notes_by_id.get(record_id, ''):
                translation_errors.append({'id': record_id, 'reason': 'missing_ai_translation_note'})
        if translation_errors:
            errors.append({'code': 'title_translation_mismatch', 'count': len(translation_errors), 'sample': translation_errors[:20]})

    report = {
        'xlsx': str(Path(args.xlsx)),
        'source_json': str(Path(args.source_json)) if args.source_json else None,
        'status': 'fail' if errors else 'pass',
        'rows': len(rows) - 1 if rows else 0,
        'alias_rows': len(alias_rows) - 1 if alias_rows else 0,
        'errors': errors,
        'warnings': warnings,
        'summary': {
            'main_rows': len(rows) - 1 if rows else 0,
            'titles_ready': len(title_index) - len(no_title_ids),
            'manual_translation_required': len(no_title_ids),
            'aliases': len(alias_rows) - 1 if alias_rows else 0,
            'alias_types': alias_type_counts,
            'alias_language_blank': alias_language_blank,
            'invalid_titles': sum(1 for item in errors if item.get('code') == 'invalid_simplified_title'),
            'untraced_source_titles': len(untraced_source_titles),
            'ai_translated_titles': len(translated_titles),
        },
        'title_blank_for_manual_translation': no_title_ids,
        'rules': {
            'title': '简体中文名只收可证实的简体中文名；外文/繁体/未确认译名进入别名明细',
            'alias': '别名明细保留原名、译名、文字变体和其他替代题名',
            'url': '入口页url不得是 sso、EZproxy、/han/ 或登录跳转',
            'unknown': '无依据字段留空并在 AI评语说明'
        }
    }
    Path(args.report).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    print(json.dumps(report, ensure_ascii=False, indent=2))
    wb.close()
    return 1 if errors else 0


if __name__ == '__main__':
    sys.exit(main())
