"""
通用脚本：扫描批次目录下 01-采集源数据/ 中所有 JSON 文件，为每个 JSON 生成 XLSX
输出到同一目录（01-采集源数据/）
每个 XLSX 包含：
  - Sheet 1: 概览（KPI 仪表盘，含数据库总量/查重后/重复数）
  - Sheet 2: 数据明细（查重后的全部字段明细表）
  - Sheet 3: 重复单元表（AI 判定的重复数据库完整数据 + 判定依据）
  - Sheet 4: 查重说明（本次查重规则与统计）
同时输出查重后 JSON（{序号}-{馆名}（{查重后数}-查重后）.json）

查重方式（第 6 步）：
  - 查重由 AI 语义比对完成，不写查重脚本。
  - AI 判定结果通过 --dup-ids 参数传入（移出主表的重复资源ID），
    或通过 --dup-file 传入包含"保留ID,重复ID,组名,判定依据"行的文本文件。

命名规则（第 6 步采集后查重）：
  - JSON 文件名标注采集总数：{序号}-{馆名}（{总数}）.json
  - Excel 文件名标注查重后数量：{序号}-{馆名}（{查重后数}-查重后）.xlsx

用法：
  python json_to_xlsx.py <批次目录路径> [--dup-ids id1,id2,...] [--dup-file path.txt]

示例：
  python json_to_xlsx.py "00-采集批次\02-公立图书馆自建+外购数据库采集（20260802）"
  python json_to_xlsx.py "00-采集批次\02-公立图书馆自建+外购数据库采集（20260802）" --dup-ids 51006,00215,00135,00117,00028,10802,10807,10810,00201,51062
"""

import json, os, sys, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== 样式常量 ==========
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E79")
META_FONT = Font(name="Arial", size=10, color="555555")
KPI_LABEL_FONT = Font(name="Arial", bold=True, size=10, color="1F4E79")
KPI_VALUE_FONT = Font(name="Arial", bold=True, size=15, color="1F4E79")
DATA_FONT = Font(name="Arial", size=9)
ZEBRA_FILL_1 = PatternFill(fill_type="solid", fgColor="FFFFFF")
ZEBRA_FILL_2 = PatternFill(fill_type="solid", fgColor="F2F7FB")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE7"),
    right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"),
    bottom=Side(style="thin", color="D9DEE7"),
)
KPI_FILL = PatternFill(fill_type="solid", fgColor="EAF2FF")

# ========== JSON 字段 → 中文列名映射 ==========
FIELD_MAP = {
    "resourceId": "资源ID",
    "name": "名称",
    "description": "描述",
    "accessMethods": "访问方式",
    "intranetUrl": "内网URL",
    "extranetUrl": "外网URL",
    "roles": "授权角色",
    "resourceType": "资源类型",
    "thumbnail": "缩略图",
    "sourcePage": "来源页",
    "url": "入口URL",
    "type": "类型",
    "subject": "学科",
    "language": "语言",
    "provider": "提供商",
    "license": "许可",
}

# 史学关键词（用于自动分类统计）
HISTORY_KEYWORDS = ["历史", "古籍", "民国", "报纸", "档案", "方志", "文献", "手稿", "拓片", "碑帖",
                    "甲骨", "敦煌", "地方志", "家谱", "族谱", "年谱", "实录", "史料", "丛刊",
                    "文史", "考古", "文物", "博物", "国学", "经典", "典籍", "石刻", "金石"]
SELF_BUILT_KEYWORDS = ["自建", "特色", "馆藏", "数字", "典藏", "数字资源", "数字化"]

# 默认查重规则说明（写入「查重说明」sheet，可随批次调整）
DEFAULT_RULES = [
    "1. 名称完全相同或入口URL完全相同（含简称，如“国研网”=“国务院发展研究中心信息网”）→ 重复",
    "2. 同一数据库的新旧入口/旧版入口、升级版/超集版（如“新版”“增补版”含旧版全部内容）→ 重复",
    "3. 同一平台同一资源、描述一致 → 重复",
    "4. 同一提供商的不同受众/版本（少儿版、公图版、基础教育版、推广工程资源等）→ 视为重复，保留主版",
    "5. 名称相近但内容不同（如博士/硕士论文库、全唐诗/全宋诗、EAI 系列 I/II 等）→ 保留为独立条目",
]


def extract_library_name(json_path):
    """从文件名提取图书馆名（去掉序号前缀、扩展名和括号数量标注）"""
    basename = os.path.splitext(os.path.basename(json_path))[0]
    basename = re.sub(r"^\d+-", "", basename)
    basename = re.sub(r"（\d+）$", "", basename)   # 去掉 （227） 数量标注
    basename = re.sub(r"（\d+-查重后）$", "", basename)
    return basename


def strip_count_mark(basename):
    """去掉文件名中的（数量）或（数量-查重后）标注，用于生成新命名"""
    return re.sub(r"（\d+-查重后）$", "", re.sub(r"（\d+）$", "", basename))


def classify_database(item):
    text = (item.get("name", "") + " " + item.get("description", "")).lower()
    is_history = any(kw in text for kw in HISTORY_KEYWORDS)
    is_self_built = any(kw in text for kw in SELF_BUILT_KEYWORDS)
    return is_history, is_self_built


def is_online(item):
    return "互联网" in item.get("accessMethods", "")


def parse_dup_args(argv):
    """解析 --dup-ids / --dup-file 参数，返回 dup_info 列表
    dup_info: [(keep_id, dup_id, group, reason), ...]"""
    dup_ids = []
    dup_info = []
    i = 0
    while i < len(argv):
        if argv[i] == "--dup-ids" and i + 1 < len(argv):
            dup_ids = [x.strip() for x in argv[i + 1].split(",") if x.strip()]
            i += 2
        elif argv[i] == "--dup-file" and i + 1 < len(argv):
            with open(argv[i + 1], encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = [p.strip() for p in line.split("|")]
                    if len(parts) >= 4:
                        dup_info.append((parts[0], parts[1], parts[2], parts[3]))
            i += 2
        else:
            i += 1
    # 若无详情，仅按 dup_ids 生成基础信息
    if dup_ids and not dup_info:
        dup_info = [(d, d, "", "") for d in dup_ids]
    return dup_info


def build_overview_sheet(ws, library_name, json_path, data, dup_count):
    """构建概览（KPI 仪表盘）Sheet"""
    ws.title = "概览"
    total = len(data) + dup_count
    online = sum(1 for d in data if is_online(d))
    offline = len(data) - online
    history_count = sum(1 for d in data if classify_database(d)[0])
    self_built_count = sum(1 for d in data if classify_database(d)[1])
    non_history = len(data) - history_count

    ws.merge_cells("B1:J1")
    c = ws["B1"]
    c.value = f"{library_name}（查重后）"
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("B2:J2")
    c = ws["B2"]
    c.value = f"数据来源: {os.path.basename(json_path)}  |  原始 {total} 条  →  查重后 {len(data)} 条（移除 {dup_count} 条重复）"
    c.font = META_FONT
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 10

    kpis = [
        ("数据库总量", total),
        ("查重后数据库", len(data)),
        ("重复数据库", dup_count),
        ("线上数据库", online),
        ("线下数据库", offline),
        ("史学数据库", history_count),
        ("非史学数据库", non_history),
        ("自建/特色库", self_built_count),
    ]
    for i, (label, value) in enumerate(kpis):
        col = i + 2
        cell = ws.cell(row=4, column=col, value=label)
        cell.font = KPI_LABEL_FONT
        cell.fill = KPI_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = KPI_VALUE_FONT
        cell.fill = KPI_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 35
    for col_idx in range(2, 10):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.cell(row=7, column=2, value="注：重复判定由 AI 语义比对（名称/入口URL/描述）完成；同一提供商不同受众/版本（少儿版、公图版等）也视为重复，保留主版。").font = Font(
        name="Arial", size=9, color="999999", italic=True
    )


def build_detail_sheet(ws, data):
    """构建数据明细 Sheet（查重后）"""
    ws.title = "数据明细"
    if not data:
        return
    all_keys = []
    seen = set()
    for d in data:
        for k in d:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)
    priority_keys = ["resourceId", "name", "description", "accessMethods",
                     "intranetUrl", "extranetUrl", "roles", "resourceType"]
    ordered_keys = priority_keys + [k for k in all_keys if k not in priority_keys]

    ws.cell(row=1, column=1, value="序号")
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).border = THIN_BORDER
    for i, key in enumerate(ordered_keys):
        col = i + 2
        cell = ws.cell(row=1, column=col, value=FIELD_MAP.get(key, key))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    for row_idx, item in enumerate(data):
        excel_row = row_idx + 2
        fill = ZEBRA_FILL_1 if row_idx % 2 == 0 else ZEBRA_FILL_2
        cell = ws.cell(row=excel_row, column=1, value=row_idx + 1)
        cell.font = DATA_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        for col_idx, key in enumerate(ordered_keys):
            val = item.get(key, "")
            if isinstance(val, list):
                val = " / ".join(str(v) for v in val)
            cell = ws.cell(row=excel_row, column=col_idx + 2, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

    col_widths = {1: 6}
    for i, key in enumerate(ordered_keys):
        if key == "description":
            col_widths[i + 2] = 60
        elif key in ("intranetUrl", "extranetUrl", "thumbnail"):
            col_widths[i + 2] = 40
        elif key in ("name", "roles"):
            col_widths[i + 2] = 30
        else:
            col_widths[i + 2] = 15
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    last_col = len(ordered_keys) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{len(data) + 1}"


def build_dup_unit_sheet(ws, data, dup_info):
    """构建重复单元表：AI 判定重复的数据库完整数据 + 判定依据"""
    ws.title = "重复单元表"
    if not dup_info:
        return
    by_id = {d.get("resourceId"): d for d in data}
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "重复单元表"
    c.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = "以下为馆内确认重复的数据库（AI 语义比对判定），已从「数据明细」移出，保留信息完整的主条目。"
    c.font = Font(name="Arial", size=9, color="999999", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    headers = ["序号", "重复组", "判定依据", "处理状态", "保留条目", "资源ID", "名称", "描述",
               "访问方式", "内网URL", "外网URL", "授权角色", "资源类型", "缩略图"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 26

    row = 4
    for idx, (keep_id, dup_id, group, reason) in enumerate(dup_info, 1):
        dup = by_id.get(dup_id, {})
        keep_name = by_id.get(keep_id, {}).get("name", "")
        fill = ZEBRA_FILL_1 if idx % 2 == 0 else ZEBRA_FILL_2
        vals = [idx, group or f"{keep_name} / {dup.get('name', '')}", reason, "移出主表（重复）", keep_name,
                dup.get("resourceId", ""), dup.get("name", ""), dup.get("description", ""),
                dup.get("accessMethods", ""), dup.get("intranetUrl", ""), dup.get("extranetUrl", ""),
                dup.get("roles", ""), dup.get("resourceType", ""), dup.get("thumbnail", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = DATA_FONT
            cell.fill = WARN_FILL if ci in (1, 4) else fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if ci in (1, 4, 6, 13) else "left",
                                       vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 60
        row += 1
    widths = [6, 30, 45, 14, 30, 10, 30, 40, 22, 32, 32, 26, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells(f"A{row}:N{row}")
    c = ws.cell(row=row, column=1)
    c.value = f"共 {len(dup_info)} 条重复数据库，已移入本表并从「数据明细」移除。"
    c.font = Font(name="Arial", size=9, color="999999", italic=True)


def build_dedup_sheet(ws, rules, dup_count, kept_count, rules_note=""):
    """构建查重说明 Sheet：本次查重规则与统计"""
    ws.title = "查重说明"
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "查重说明"
    c.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "本次查重由 AI 语义比对完成，规则如下："
    c.font = Font(name="Arial", size=9, color="999999", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22
    row = 3
    for r in rules:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=r)
        c.font = Font(name="Arial", size=9, color="333333")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22
        row += 1
    if rules_note:
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=rules_note)
        c.font = Font(name="Arial", size=9, color="333333")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1)
    c.value = f"重复 {dup_count} 条，已移出；查重后 {kept_count} 条。"
    c.font = Font(name="Arial", size=9, color="999999", italic=True)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40


def convert_json_to_xlsx(json_path, output_dir, dup_info, rename_json=False, rules=None, rules_note=""):
    """将单个 JSON 文件转换为 XLSX（按 AI 查重结果），并按数量规则命名"""
    basename = os.path.splitext(os.path.basename(json_path))[0]
    library_name = extract_library_name(json_path)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        data = [data]
    total = len(data)

    dup_ids = {dup_id for _, dup_id, _, _ in dup_info}
    kept_data = [d for d in data if d.get("resourceId") not in dup_ids]

    # 重命名 JSON：{序号}-{馆名}（{总数}）.json
    if rename_json:
        clean_base = strip_count_mark(basename)
        new_json_name = f"{clean_base}（{total}）.json"
        new_json_path = os.path.join(output_dir, new_json_name)
        if os.path.abspath(new_json_path) != os.path.abspath(json_path):
            os.rename(json_path, new_json_path)
            json_path = new_json_path
            basename = os.path.splitext(os.path.basename(json_path))[0]

    clean_base = strip_count_mark(basename)

    # 生成查重后 JSON：{序号}-{馆名}（{查重后数}-查重后）.json
    kept_json_name = f"{clean_base}（{len(kept_data)}-查重后）.json"
    kept_json_path = os.path.join(output_dir, kept_json_name)
    with open(kept_json_path, "w", encoding="utf-8") as f:
        json.dump(kept_data, f, ensure_ascii=False, indent=2)

    # 生成 XLSX：{序号}-{馆名}（{查重后数}-查重后）.xlsx
    xlsx_name = f"{clean_base}（{len(kept_data)}-查重后）.xlsx"
    output_path = os.path.join(output_dir, xlsx_name)

    wb = Workbook()
    build_overview_sheet(wb.active, library_name, json_path, kept_data, len(dup_info))
    build_detail_sheet(wb.create_sheet(), kept_data)
    build_dup_unit_sheet(wb.create_sheet(), data, dup_info)
    build_dedup_sheet(wb.create_sheet(), rules or DEFAULT_RULES, len(dup_info), len(kept_data), rules_note)
    wb.save(output_path)
    return output_path, total, len(kept_data), len(dup_info)


def process_batch(batch_dir, dup_info):
    """处理整个批次目录"""
    src_data_dir = os.path.join(batch_dir, "01-采集源数据")
    os.makedirs(src_data_dir, exist_ok=True)

    results = []
    for fname in sorted(os.listdir(src_data_dir)):
        if not fname.endswith(".json"):
            continue
        if "_to_wordpress" in fname or ".progress" in fname:
            continue
        json_path = os.path.join(src_data_dir, fname)
        try:
            output_path, total, kept, dup = convert_json_to_xlsx(json_path, src_data_dir, dup_info, rename_json=True)
            results.append((fname, total, kept, dup, "成功"))
            print(f"  ✓ {fname} → {os.path.basename(output_path)}（原始 {total} / 查重后 {kept} / 重复 {dup}）")
        except Exception as e:
            results.append((fname, 0, 0, 0, f"失败: {e}"))
            print(f"  ✗ {fname} 失败: {e}")

    print(f"\n共处理 {len(results)} 个 JSON 文件")
    return results


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python json_to_xlsx.py <批次目录路径> [--dup-ids id1,id2,...] [--dup-file path.txt]")
        sys.exit(1)
    batch_dir = sys.argv[1]
    dup_info = parse_dup_args(sys.argv[2:])
    if dup_info:
        print(f"AI 判定重复 {len(dup_info)} 条: {[d[1] for d in dup_info]}")
    else:
        print("提示: 未提供 --dup-ids/--dup-file，AI 查重结果请通过参数传入；当前仅生成无查重的汇总表。")
    print(f"处理批次目录: {batch_dir}")
    process_batch(batch_dir, dup_info)
