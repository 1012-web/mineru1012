# -*- coding: utf-8 -*-
"""批次总目生成（规范模板 v2，47 列）

把某个馆的采集结果填进批次总目模板的「条目表」。

设计前提：**采集阶段不受约束**。各馆按自己的原貌采，字段叫什么、嵌在第几层，
都由该馆的源站决定。所以本脚本不认死任何字段名，而是读两个外部 JSON：

  --field-map   这个馆的字段叫什么（结构适配，见 references/映射-*.json）
  --judgments   这个馆的 AI 语义判断结果（主办方/网址类型/自建/史学…，见 references/实例-*判断.json）

换一个馆＝换这两个文件，脚本本身一行不动。

用法：

    python scripts/build_batch_mulu.py \
        --json  "00-采集批次/02-.../01-采集源数据/01-国家图书馆（217-查重后）.json" \
        --out   "00-采集批次/02-.../批次总目.xlsx" \
        --field-map "references/映射-国家图书馆.json" \
        --judgments "references/实例-国家图书馆判断.json" \
        --batch-name "02-公立图书馆自建+外购数据库采集（20260802）" \
        --source-name "国家图书馆" \
        --source-url  "http://read.nlc.cn/outRes/outResList?type=全部" \
        --check-date 2026-08-02 \
        --id-prefix NLC \
        --sso-host sso1.nlc.cn

多馆批次：第一个馆正常跑，之后每个馆加 --append，条目累加到同一个批次总目。
没有 --append 会清空既有行 —— 多馆批次漏加就会把前面的馆抹掉。

--field-map 的结构（全部可缺省）：

    {
      "records_path": "databases",       记录数组在哪；null＝顶层就是数组；
                                         支持 "resourceGroups.databases" 这种嵌套（逐层展平）
      "id_fallback": "index",            源站没有 id 字段时按序号生成
      "fields": {                        左边是脚本要用的语义名，右边是该馆的实际字段名
        "id": "resourceId",
        "name": "name",
        "description": "description",
        "entry_url": "intranetUrl",
        "external_url": "extranetUrl",
        "source_page": "sourcePage"
      },
      "access": {                        怎么判断馆内/馆外访问
        "field": "accessMethods",
        "inlibrary_contains": ["局域网"],   也支持 inlibrary_equals / offsite_equals
        "offsite_contains": ["互联网"]      两边都不命中 → 未说明，并写进 AI 评语
      },
      "remark_tags": [                   备注列的原始标签，原样留存不翻译不归并
        {"label": "访问方式", "key": "accessMethods"}
      ]
    }
"""
import argparse
import html
import json
import os
import sys

from openpyxl import load_workbook

TOTAL_COLS = 47

COLUMN_LABELS = {
    'id': '条目ID', 'batch': '采集批次', 'src_name': '来源名称', 'src_url': '来源URL',
    'status': '复核状态', 'self': '自建/特色库？', 'hist': '史学数据库？',
    'name': '简体中文名', 'stype': '网址类型', 'carrier': '载体', 'org': '主办方',
    'url': '入口页url', 'access': '访问方式', 'ui_lang': '网站语言',
    'country': '网站国家/地区', 'auth_org': '授权机构（本批次）',
    'lib_access': '该馆访问方式', 'local_name': '机构内名称',
    'access_page': '授权页url', 'check_date': '核验日期', 'brief': '简介',
    'db_type': '数据库类型',
    'score': 'AI评分', 'note': 'AI评语', 'remark': '备注',
}
# 图书馆导航页上的介绍是图书馆自己缩写的，不是数据库官网原文，所以进「简介」不进「官网介绍」。
# 「官网介绍」按定义只收数据库自身官网的完整连续原文，本环节留空，编目阶段再抓。


def clean_html(s):
    return html.unescape(s) if s else ''


# ---------- 结构适配 ----------

def load_field_map(path):
    empty = {'records_path': None, 'id_fallback': 'index', 'fields': {},
             'access': {}, 'remark_tags': []}
    if not path:
        return empty
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    out = dict(empty)
    for k in empty:
        if k in raw:
            out[k] = raw[k]
    return out


def extract_records(data, records_path):
    """按 records_path 取出记录数组。支持 'a' 和 'a.b'（逐层展平）。"""
    if records_path in (None, '', False):
        if isinstance(data, list):
            return data
        sys.exit('采集 JSON 顶层不是数组，请在 --field-map 里用 records_path 指明记录在哪个键下')
    cur = data
    for part in str(records_path).split('.'):
        if isinstance(cur, list):
            nxt = []
            for item in cur:
                got = item.get(part) if isinstance(item, dict) else None
                if isinstance(got, list):
                    nxt.extend(got)
            cur = nxt
        elif isinstance(cur, dict):
            cur = cur.get(part)
        else:
            cur = None
        if cur is None:
            sys.exit(f'按 records_path="{records_path}" 取不到记录，卡在 "{part}"')
    if not isinstance(cur, list):
        sys.exit(f'records_path="{records_path}" 指向的不是数组，而是 {type(cur).__name__}')
    return cur


def get_field(rec, fmap, semantic, default=''):
    key = fmap['fields'].get(semantic)
    if not key:
        return default
    val = rec.get(key, default)
    return default if val is None else val


def resolve_access(rec, fmap):
    """返回 (仅馆内?, 可馆外?, 判定成功?)。字段对不上就都是 False，由调用方记进评语。"""
    cfg = fmap.get('access') or {}
    field = cfg.get('field')
    if not field or field not in rec:
        return False, False, False
    val = rec.get(field)
    text = '' if val is None else str(val)

    def hit(kind):
        for needle in cfg.get(f'{kind}_contains', []):
            if str(needle) in text:
                return True
        for exact in cfg.get(f'{kind}_equals', []):
            if val == exact or text == str(exact):
                return True
        return False

    inlib, offsite = hit('inlibrary'), hit('offsite')
    return inlib, offsite, (inlib or offsite)


# ---------- AI 判断结果 ----------

def load_judgments(path):
    empty = {
        'organizer': {}, 'site_type': {}, 'ui_lang': {}, 'country': {},
        'db_type': {}, 'org_notes': {}, 'type_notes': {},
        'self_built_yes': set(), 'self_built_doubt': set(),
        'history_yes': set(), 'history_doubt': set(),
    }
    if not path:
        return empty
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    out = dict(empty)
    for k in ('organizer', 'site_type', 'ui_lang', 'country', 'db_type',
              'org_notes', 'type_notes'):
        out[k] = raw.get(k, {})
    for k in ('self_built_yes', 'self_built_doubt', 'history_yes', 'history_doubt'):
        out[k] = set(raw.get(k, []))
    return out


# ---------- 工作簿 ----------

def build_column_map(ws):
    headers = [ws.cell(row=1, column=c).value for c in range(1, TOTAL_COLS + 1)]
    col = {h: i for i, h in enumerate(headers, 1)}
    missing = [label for label in COLUMN_LABELS.values() if label not in col]
    if missing:
        sys.exit(f'模板「条目表」缺少这些列，无法写入：{"、".join(missing)}')
    return {k: col[label] for k, label in COLUMN_LABELS.items()}


def first_empty_row(ws):
    for r in range(2, ws.max_row + 2):
        if all(ws.cell(row=r, column=c).value in (None, '') for c in range(1, TOTAL_COLS + 1)):
            return r
    return ws.max_row + 1


def clear_rows(ws):
    for r in range(2, ws.max_row + 1):
        for c in range(1, TOTAL_COLS + 1):
            ws.cell(row=r, column=c).value = None


def main():
    p = argparse.ArgumentParser(description='把某个馆的采集结果填进批次总目规范模板')
    p.add_argument('--json', required=True, help='该馆的采集 JSON（建议用查重后的）')
    p.add_argument('--out', required=True, help='输出的批次总目 xlsx')
    p.add_argument('--field-map', help='该馆的字段对应表；不给则默认顶层数组 + 无字段可取')
    p.add_argument('--judgments', help='该馆的 AI 判断结果；不给则全部走默认值')
    p.add_argument('--template', default=r'D:\projects\目录1012题录库\批次总目模板.xlsx',
                   help='规范模板；--append 时改用 --out 已有文件')
    p.add_argument('--batch-name', required=True, help='采集批次列的值')
    p.add_argument('--source-name', required=True, help='来源名称，通常是馆名')
    p.add_argument('--source-url', required=True,
                   help='该馆电子资源导航页 URL；同时用作授权页url 与官网介绍出处url')
    p.add_argument('--check-date', required=True, help='核验日期 YYYY-MM-DD')
    p.add_argument('--id-prefix', required=True, help='条目ID 前缀，如 NLC → NLC-001')
    p.add_argument('--auth-org', help='授权机构（本批次）；缺省＝--source-name')
    p.add_argument('--tag-prefix', help='备注列原始标签的来源前缀；缺省＝--source-name')
    p.add_argument('--sso-host', help='代理跳转域名，如 sso1.nlc.cn；命中则在评语注明未收录')
    p.add_argument('--expect', type=int, help='期望条数；不符即报错退出，防止拿错文件')
    p.add_argument('--append', action='store_true', help='追加到 --out 已有的批次总目，不清空既有行')
    a = p.parse_args()

    auth_org = a.auth_org or a.source_name
    tag_prefix = a.tag_prefix or a.source_name

    fmap = load_field_map(a.field_map)
    j = load_judgments(a.judgments)

    with open(a.json, encoding='utf-8') as f:
        data = json.load(f)
    records = extract_records(data, fmap.get('records_path'))
    if a.expect is not None and len(records) != a.expect:
        sys.exit(f'条数不符：{a.json} 实际 {len(records)} 条，--expect 要求 {a.expect} 条')
    if not records:
        sys.exit(f'没有取到任何记录：{a.json}（该馆可能采集失败，先看采集产物再填充）')

    if a.append:
        if not os.path.exists(a.out):
            sys.exit(f'--append 需要 --out 已存在：{a.out}')
        wb = load_workbook(a.out)
    else:
        wb = load_workbook(a.template)
    if '条目表' not in wb.sheetnames:
        sys.exit(f'模板里没有「条目表」这个 sheet，实际有：{"、".join(wb.sheetnames)}')
    ws = wb['条目表']

    C = build_column_map(ws)
    if a.append:
        start = first_empty_row(ws)
    else:
        clear_rows(ws)
        start = 2

    unresolved_access = 0
    for i, d in enumerate(records):
        r = start + i
        notes = []

        rid = get_field(d, fmap, 'id')
        if not rid and fmap.get('id_fallback') == 'index':
            rid = str(i + 1)
        rid = str(rid)

        name = clean_html(str(get_field(d, fmap, 'name')))
        desc = clean_html(str(get_field(d, fmap, 'description')))

        inlib, offsite, resolved = resolve_access(d, fmap)
        if not resolved:
            unresolved_access += 1
            notes.append('访问方式字段对不上，填未说明待核')
        parts = []
        if inlib:
            parts.append('仅馆内访问')
        if offsite:
            parts.append('可馆外访问')
        access_main = '需授权＞机构订购' + ('、' + '、'.join(parts) if parts else '')
        lib_access = '可馆外' if offsite else ('仅馆内' if inlib else '未说明')

        ws.cell(row=r, column=C['id'], value=f'{a.id_prefix}-{i + 1:03d}')
        ws.cell(row=r, column=C['batch'], value=a.batch_name)
        ws.cell(row=r, column=C['src_name'], value=a.source_name)
        ws.cell(row=r, column=C['src_url'], value=a.source_url)
        ws.cell(row=r, column=C['status'], value='待复核')

        if rid in j['self_built_yes']:
            self_built = '是'
        elif rid in j['self_built_doubt']:
            self_built = '存疑'
        else:
            self_built = '否'
        ws.cell(row=r, column=C['self'], value=self_built)

        if rid in j['history_yes']:
            hist = '是'
        elif rid in j['history_doubt']:
            hist = '存疑'
        else:
            hist = '否'
        ws.cell(row=r, column=C['hist'], value=hist)

        ws.cell(row=r, column=C['name'], value=name)
        ws.cell(row=r, column=C['stype'], value=j['site_type'].get(rid, '数据库'))
        ws.cell(row=r, column=C['carrier'], value='网页')
        ws.cell(row=r, column=C['org'], value=j['organizer'].get(rid, ''))
        ws.cell(row=r, column=C['url'], value=get_field(d, fmap, 'entry_url'))
        ws.cell(row=r, column=C['access'], value=access_main)
        ws.cell(row=r, column=C['ui_lang'], value=j['ui_lang'].get(rid, '简体中文'))
        ws.cell(row=r, column=C['country'], value=j['country'].get(rid, '中国'))

        # 按馆采集五列
        ws.cell(row=r, column=C['auth_org'], value=auth_org)
        ws.cell(row=r, column=C['lib_access'], value=lib_access)
        ws.cell(row=r, column=C['local_name'], value=name)
        ws.cell(row=r, column=C['access_page'], value=a.source_url)
        ws.cell(row=r, column=C['check_date'], value=a.check_date)

        ws.cell(row=r, column=C['brief'], value=desc)
        ws.cell(row=r, column=C['db_type'], value=j['db_type'].get(rid, ''))

        score = 90
        if rid not in j['organizer']:
            score = 75
            notes.append(j['org_notes'].get(rid, '主办方信息不足，留空待核'))
        if j['site_type'].get(rid):
            notes.append(j['type_notes'].get(rid, f'网址类型按决策树归{j["site_type"][rid]}'))
        if rid in j['self_built_doubt']:
            notes.append('自建/特色库存疑，待人工复核')
        if rid in j['history_doubt']:
            notes.append('史学数据库存疑，待人工复核')
        if not desc:
            notes.append('馆方无介绍文本，简介留空')
        if a.sso_host and a.sso_host in str(get_field(d, fmap, 'external_url')):
            notes.append('入口页url取内网地址；sso代理链接未收录')
        ws.cell(row=r, column=C['score'], value=score)
        ws.cell(row=r, column=C['note'], value='；'.join(notes))

        tags = [f'{tag_prefix}｜{t["label"]}：{d.get(t["key"], "")}'
                for t in fmap.get('remark_tags', [])]
        ws.cell(row=r, column=C['remark'], value='\n'.join(tags) if tags else None)

    wb.save(a.out)
    verb = '追加' if a.append else '写入'
    print(f'{verb} {len(records)} 条（第 {start} 行起）-> {a.out}')
    if unresolved_access:
        print(f'⚠ {unresolved_access} 条的访问方式判不出来，已填「未说明」并写进 AI 评语。'
              f'检查 --field-map 的 access 配置是否匹配该馆字段。')


if __name__ == '__main__':
    main()
