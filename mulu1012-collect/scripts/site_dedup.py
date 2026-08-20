# -*- coding: utf-8 -*-
"""入库查重 · 脚本粗筛 + 人工确认表

三层漏斗的机器部分：脚本把「必然重复」和「必然不重复」分掉，AI 只判灰区，用户拍板。

数据放两处，职责分开：

    02-查重/查重结论-{馆名}.jsonl     真源。一行一条，可 diff、可追加、可跨批次拼接。
    批次总目.xlsx 的「站内查重」sheet  渲染产物 + 人工输入口。

**跑一次命令 = 回收人工列 → 更新结论 → 重新渲染**，顺序固定，跑几次都安全。
`by` 字段决定谁能覆盖谁：user > ai > script。你在表里填过的，AI 和脚本都不会动。

典型用法：

    # 1. 粗筛某个馆（首次，或采集数据更新后重跑）
    python scripts/site_dedup.py \
        --json "00-采集批次/02-.../01-采集源数据/01-国家图书馆（227）.json" \
        --field-map "references/映射-国家图书馆.json" \
        --source-name "国家图书馆" \
        --dedup-dir "00-采集批次/02-.../02-查重" \
        --xlsx "00-采集批次/02-.../批次总目.xlsx"

    # 2. AI 二审：直接编辑 jsonl，给「疑似」档填 verdict/confidence/reason，by 写 "ai"

    # 3. 只渲染（不重跑粗筛）——AI 审完后刷新表格给人看
    python scripts/site_dedup.py \
        --dedup-dir "00-采集批次/02-.../02-查重" \
        --xlsx "00-采集批次/02-.../批次总目.xlsx"

    # 4. 你在「站内查重」sheet 的「人工判定」列填改判，再跑一次步骤 3 即回收

JSONL 每行结构：

    {"id": "00154", "source": "国家图书馆",
     "name": "中国重要报纸全文数据库（清华同方知网）",
     "tier": "疑似",                     脚本粗筛档：已存在/疑似/未匹配
     "hits": [{"wpid": 12175, "name": "...", "why": "名称包含：..."}],
     "verdict": "已存在",                 最终结论（五类，见 SKILL.md §2.4）
     "wpid": 12175,                       判已存在时指向的站内条目
     "confidence": 95,                    0-100 整数
     "reason": "...",                     一句话依据，给半年后的自己看
     "by": "ai",                          script / ai / user
     "at": "2026-08-02",
     "note": ""}                          人工备注
"""
import argparse
import datetime
import difflib
import glob
import json
import os
import re
import sys
import unicodedata
from urllib.parse import urlsplit

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_CATALOG = r'D:\projects\目录1012题录库\总目-与站点同步.xlsx'
SHEET = '站内查重'
VERDICTS = ['已存在', '疑似重复', '母库已存在子库未单列', '未收录', '证据不足']
TIER_ORDER = {'疑似': 0, '已存在': 1, '未匹配': 2}

# 列的取舍依据：人工判定重复与否靠「名字 + 主办者 + 链接」三样，
# 两边的这三样都要在同一行看得见，否则人得另开浏览器查，审不动。
COLUMNS = [
    ('条目ID', 11), ('来源馆', 24), ('粗筛档', 8),
    ('采集名称', 34), ('采集入口', 34), ('采集简介', 50),
    ('候选（站内）', 34), ('候选主办方', 22), ('候选入口', 34), ('匹配依据', 30),
    ('AI判定', 13), ('把握', 6), ('AI理由', 40),
    ('人工判定', 13), ('人工WPID', 10), ('人工备注', 26),
]
MANUAL_COLS = ('人工判定', '人工WPID', '人工备注')
# 简介判重时用不上（靠名字/主办者/链接就够），但偶尔要翻——留着列、默认折叠起来。
HIDDEN_COLS = ('采集简介',)
# 判重时眼睛主要在这两列来回跳，给它们更大字号、加粗、独立底色。
NAME_COLS = ('采集名称', '候选（站内）')

TIER_FILL = {'疑似': 'FFF4CE', '已存在': 'E2F0D9', '未匹配': 'F2F2F2'}
VERDICT_FILL = {'已存在': 'E2F0D9', '疑似重复': 'FFF4CE',
                '母库已存在子库未单列': 'FFE8D6', '未收录': 'DEEBF7', '证据不足': 'F2F2F2'}

# ---------- 规范化 ----------

_PUNCT = re.compile(r'[\s·・••\-—_–:：;；,，.。()（）\[\]【】《》<>「」『』"\'""''!！?？/\\|]+')


def norm_name(s):
    if not s:
        return ''
    return _PUNCT.sub('', unicodedata.normalize('NFKC', str(s)).lower())


def norm_url(u):
    if not u:
        return ''
    u = str(u).strip()
    if not u or u in ('#', '-'):
        return ''
    try:
        parts = urlsplit(u if '://' in u else 'http://' + u)
    except ValueError:
        return u.lower()
    host = (parts.netloc or '').lower()
    if host.startswith('www.'):
        host = host[4:]
    q = ('?' + parts.query) if parts.query else ''
    return (host + (parts.path or '').rstrip('/') + q).lower()


def url_host(u):
    n = norm_url(u)
    return n.split('/')[0].split('?')[0] if n else ''


def split_aliases(s):
    return [a.strip() for a in re.split(r'[;；、,，/|]', str(s or '')) if a.strip()]


# ---------- 读总目 ----------

def load_catalog(path):
    wb = load_workbook(path, read_only=True)
    if '网址总表' not in wb.sheetnames:
        sys.exit(f'{path} 里没有「网址总表」，实际有：{"、".join(wb.sheetnames)}')
    ws = wb['网址总表']
    rows = ws.iter_rows(values_only=True)
    header = [h if h else '' for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    for need in ('WPID', '简体中文名'):
        if need not in idx:
            sys.exit(f'网址总表缺少「{need}」列')

    entries = {}
    for r in rows:
        def get(col):
            i = idx.get(col)
            return r[i] if i is not None and i < len(r) else None
        name = get('简体中文名')
        if not name:
            continue
        wpid = get('WPID')
        key = str(wpid) if wpid else f'row:{len(entries)}'
        entry_url = get('入口页url')
        entries[key] = {
            'wpid': wpid,
            'name': str(name),
            'org': str(get('主办方') or ''),          # 人工判重三要素之一
            'entry': str(entry_url or ''),           # 原样保留，给人看；匹配用规范化后的
            'aliases': split_aliases(get('别名')),
            'urls': [u for u in (norm_url(entry_url), norm_url(get('介绍页url'))) if u],
        }

    if '别名明细' in wb.sheetnames:
        rows2 = wb['别名明细'].iter_rows(values_only=True)
        header2 = [h if h else '' for h in next(rows2)]
        i2 = {h: i for i, h in enumerate(header2)}
        if 'WPID' in i2 and '名称' in i2:
            for r in rows2:
                wpid, alias = r[i2['WPID']], r[i2['名称']]
                if wpid is not None and alias and str(wpid) in entries:
                    entries[str(wpid)]['aliases'].append(str(alias))
    wb.close()
    return list(entries.values())


# ---------- 读采集 JSON ----------

def extract_records(data, records_path):
    if records_path in (None, '', False):
        if isinstance(data, list):
            return data
        sys.exit('采集 JSON 顶层不是数组，请在 field-map 用 records_path 指明记录在哪个键下')
    cur = data
    for part in str(records_path).split('.'):
        if isinstance(cur, list):
            nxt = []
            for item in cur:
                got = item.get(part) if isinstance(item, dict) else None
                if isinstance(got, list):
                    nxt.extend(got)
            cur = nxt
        elif isinstance(cur, dict):
            cur = cur.get(part)
        if cur is None:
            sys.exit(f'按 records_path="{records_path}" 取不到记录，卡在 "{part}"')
    return cur if isinstance(cur, list) else []


# ---------- 匹配 ----------

def as_hit(e, why):
    """候选带上主办方和入口——人工判重靠名字+主办者+链接，缺一样就得另开浏览器。"""
    return {'wpid': e['wpid'], 'name': e['name'], 'org': e.get('org', ''),
            'entry': e.get('entry', ''), 'why': why}


def match_one(name, urls, catalog, name_index, url_index, sim):
    n = norm_name(name)
    hits = []

    for u in urls:
        for e in url_index.get(u, []):
            hits.append(as_hit(e, f'入口URL相同：{u}'))
    if hits:
        return '已存在', hits

    for e in name_index.get(n, []) if n else []:
        hits.append(as_hit(e, '名称规范化后相同'))
    if hits:
        return '已存在', hits

    for e in catalog:
        for cand in [e['name']] + e['aliases']:
            cn = norm_name(cand)
            if not cn or not n:
                continue
            why = None
            shorter = min(n, cn, key=len)
            # 超短词的包含匹配全是噪音（总目别名里有「资料」「DH」「BASE」这类）
            if len(shorter) >= 4 and (not shorter.isascii() or len(shorter) >= 10) \
                    and (n in cn or cn in n):
                why = f'名称包含：{cand}'
            elif abs(len(cn) - len(n)) <= max(len(n), len(cn)) * 0.5:
                ratio = difflib.SequenceMatcher(None, n, cn).ratio()
                if ratio >= sim:
                    why = f'名称相似 {ratio:.0%}：{cand}'
            if why:
                hits.append(as_hit(e, why))
                break
    if hits:
        return '疑似', hits

    hosts = {url_host(u) for u in urls if u}
    for e in catalog:
        for u in e['urls']:
            if (h := u.split('/')[0].split('?')[0]) and h in hosts:
                hits.append(as_hit(e, f'入口同域名：{h}'))
                break
        if len(hits) >= 5:
            break
    return ('疑似', hits) if hits else ('未匹配', [])


# ---------- JSONL 读写 ----------

def check_freshness(catalog_path, catalog_count, dedup_dir, timeout=8):
    """快照过时会把已上线的条目误判成「未收录」，导致重复上传。

    线上 sites 混着公众号/小程序，条数跟网址总表对不上，所以不比绝对值，
    只比「跟上次跑相比线上有没有变多」——基线存在查重目录里。
    网络不通就静默跳过，查重本身不该被这个卡住。
    """
    import urllib.request
    base_file = os.path.join(dedup_dir, '.线上基线.json')
    try:
        req = urllib.request.Request(
            'https://mulu.1012.wiki/wp-json/wp/v2/sites?per_page=1',
            method='HEAD', headers={'User-Agent': 'mulu1012-dedup'})
        with urllib.request.urlopen(req, timeout=timeout) as r:
            online = int(r.headers.get('X-WP-Total') or 0)
    except Exception:
        return                                  # 离线/超时都不打扰
    if not online:
        return

    mtime = datetime.date.fromtimestamp(os.path.getmtime(catalog_path)).isoformat()
    prev = {}
    if os.path.exists(base_file):
        try:
            with open(base_file, encoding='utf-8') as f:
                prev = json.load(f)
        except Exception:
            prev = {}

    if prev.get('online') and online > prev['online']:
        print(f'⚠ 线上 sites 从 {prev["online"]} 增至 {online} 条，'
              f'而快照还是 {mtime} 导出的（{catalog_count} 条网址）。')
        print('  新上线的条目会被误判成「未收录」→ 重复上传。'
              '重新导出 总目-与站点同步.xlsx 再跑。')

    os.makedirs(dedup_dir, exist_ok=True)
    with open(base_file, 'w', encoding='utf-8') as f:
        json.dump({'online': online, 'catalog_count': catalog_count,
                   'catalog_mtime': mtime, 'checked_at': datetime.date.today().isoformat()},
                  f, ensure_ascii=False, indent=1)


def jsonl_path(dedup_dir, source_name):
    return os.path.join(dedup_dir, f'查重结论-{source_name}.jsonl')


def read_jsonl(path):
    if not os.path.exists(path):
        return {}
    out = {}
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                rec = json.loads(line)
                out[rec['id']] = rec
    return out


def write_jsonl(path, records):
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + '\n')


def read_all_jsonl(dedup_dir):
    """读该批次所有馆的结论，返回 {(source, id): rec}。"""
    out = {}
    for p in sorted(glob.glob(os.path.join(dedup_dir, '查重结论-*.jsonl'))):
        for rec in read_jsonl(p).values():
            out[(rec.get('source', ''), rec['id'])] = rec
    return out


# ---------- xlsx 人工列回收 ----------

def collect_manual(xlsx, dedup_dir):
    """把「站内查重」sheet 里的人工列读回各馆 jsonl。返回回收条数。"""
    if not os.path.exists(xlsx):
        return 0
    wb = load_workbook(xlsx)
    if SHEET not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[SHEET]
    head = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    if not all(k in head for k in ('条目ID', '来源馆', *MANUAL_COLS)):
        wb.close()
        return 0

    # 条目ID → 它落在哪个 jsonl。以前是拿「来源馆」当文件名去找，那要求这一列的值
    # 必须等于 --source-name；合并型批次里这一列会写成馆名清单，于是找不到文件、
    # 人工改判被静默丢弃。改成按 id 建索引，「来源馆」怎么写都不影响回收。
    id2file = {}
    for p in sorted(glob.glob(os.path.join(dedup_dir, '查重结论-*.jsonl'))):
        for rid in read_jsonl(p):
            id2file.setdefault(str(rid), p)

    today = datetime.date.today().isoformat()
    by_source = {}
    n = 0
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, head['条目ID']).value
        src = ws.cell(r, head['来源馆']).value
        if not rid:
            continue

        def cell(label):
            v = ws.cell(r, head[label]).value
            return str(v).strip() if v not in (None, '') else ''

        verdict, wpid_in, note = cell('人工判定'), cell('人工WPID'), cell('人工备注')
        if not verdict and not wpid_in and not note:
            continue          # 三列全空＝同意 AI 判定，不动
        p = id2file.get(str(rid)) or jsonl_path(dedup_dir, str(src))
        by_source.setdefault(p, {})[str(rid)] = (verdict, wpid_in, note)
        n += 1
    wb.close()

    for p, edits in by_source.items():
        recs = read_jsonl(p)
        for rid, (verdict, wpid_in, note) in edits.items():
            rec = recs.get(rid)
            if not rec:
                continue
            if verdict:
                if verdict != '同意':
                    rec['verdict'] = verdict
                    rec['confidence'] = 100
                rec['by'] = 'user'
                rec['at'] = today
            if note:
                rec['note'] = note

            # WPID：人工指定优先；判「已存在」但没指定就取第一个候选；判非已存在则清空
            final = rec.get('verdict', '')
            if wpid_in:
                try:
                    rec['wpid'] = int(wpid_in)
                except ValueError:
                    rec['wpid'] = wpid_in
            elif final == '已存在' and not rec.get('wpid'):
                hits = rec.get('hits') or []
                rec['wpid'] = hits[0].get('wpid') if hits else None
            elif final and final != '已存在':
                rec['wpid'] = None
        if recs:
            write_jsonl(p, list(recs.values()))
    return n


# ---------- xlsx 渲染 ----------

def render(xlsx, dedup_dir):
    recs = list(read_all_jsonl(dedup_dir).values())
    if not recs:
        print('没有查重结论可渲染')
        return 0
    recs.sort(key=lambda r: (TIER_ORDER.get(r.get('tier'), 9), r.get('source', ''), r['id']))

    wb = load_workbook(xlsx)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_name_fill = PatternFill('solid', fgColor='0F3355')      # 名称列表头更深，一眼定位
    hdr_man_fill = PatternFill('solid', fgColor='7F6000')       # 人工列表头暖色，跟只读区分开
    man_fill = PatternFill('solid', fgColor='FFF2CC')
    name_fill = PatternFill('solid', fgColor='FAFAFA')
    thin = Side(style='thin', color='D0D0D0')
    edge = Side(style='medium', color='1F4E79')

    for i, (label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(1, i, label)
        c.font = Font(bold=True, color='FFFFFF', size=12 if label in NAME_COLS else 10)
        c.fill = (hdr_name_fill if label in NAME_COLS else
                  hdr_man_fill if label in MANUAL_COLS else hdr_fill)
        c.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        dim = ws.column_dimensions[get_column_letter(i)]
        dim.width = width
        if label in HIDDEN_COLS:
            dim.hidden = True
    ws.row_dimensions[1].height = 30

    wrap = ('来源馆', '采集名称', '采集入口', '采集简介', '候选（站内）', '候选主办方',
            '候选入口', '匹配依据', 'AI理由', '人工备注')
    idx = {label: i for i, (label, _) in enumerate(COLUMNS, 1)}
    for r, rec in enumerate(recs, 2):
        hits = (rec.get('hits') or [])[:3]     # 疑似档可能有多个候选，都要让人看见
        is_user = rec.get('by') == 'user'
        row = {
            # libs 是合并型批次用的：一条题录被多个馆持有时，这里写全部持有馆，
            # 单馆批次没这个键，照旧显示 --source-name。
            '条目ID': rec['id'], '来源馆': rec.get('libs') or rec.get('source', ''),
            '粗筛档': rec.get('tier', ''),
            '采集名称': rec.get('name', ''), '采集入口': rec.get('entry', ''),
            '采集简介': rec.get('brief', ''),
            '候选（站内）': '\n'.join(f'{h.get("wpid")}｜{h.get("name", "")}' for h in hits),
            '候选主办方': '\n'.join(h.get('org', '') or '—' for h in hits),
            '候选入口': '\n'.join(h.get('entry', '') or '—' for h in hits),
            '匹配依据': '\n'.join(h.get('why', '') for h in hits),
            'AI判定': '' if is_user else rec.get('verdict', ''),
            '把握': '' if is_user else rec.get('confidence', ''),
            'AI理由': '' if is_user else rec.get('reason', ''),
            '人工判定': rec.get('verdict', '') if is_user else '',
            '人工WPID': rec.get('wpid', '') if is_user else '',
            '人工备注': rec.get('note', ''),
        }
        for label, col in idx.items():
            cell = ws.cell(r, col, row[label])
            cell.alignment = Alignment(vertical='top', wrap_text=label in wrap)
            cell.border = Border(left=edge if label in NAME_COLS else thin,
                                 right=thin, top=thin, bottom=thin)
            if label in NAME_COLS:
                # 判重时眼睛在这两列来回跳：加粗、放大、浅底，扫得快
                cell.font = Font(bold=True, size=12)
                cell.fill = name_fill
            elif label in MANUAL_COLS:
                cell.fill = man_fill
            elif label == '粗筛档':
                cell.alignment = Alignment(vertical='top', horizontal='center')
                if c := TIER_FILL.get(row[label]):
                    cell.fill = PatternFill('solid', fgColor=c)
            elif label == 'AI判定':
                if c := VERDICT_FILL.get(row[label]):
                    cell.fill = PatternFill('solid', fgColor=c)
            elif label == '把握':
                cell.alignment = Alignment(vertical='top', horizontal='center')
                if isinstance(row[label], int) and row[label] < 70:
                    cell.font = Font(color='C00000', bold=True)   # 把握低的标红，优先看
        ws.row_dimensions[r].height = None      # 交给 Excel 按内容自适应

    dv = DataValidation(type='list', formula1='"' + ','.join(['同意'] + VERDICTS) + '"', allow_blank=True)
    dv.error = '请从下拉列表中选择'
    ws.add_data_validation(dv)
    dv.add(f'{get_column_letter(idx["人工判定"])}2:{get_column_letter(idx["人工判定"])}{len(recs) + 1}')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{len(recs) + 1}'
    wb.save(xlsx)
    return len(recs)


# ---------- 主流程 ----------

def screen(a, dedup_dir):
    """粗筛一个馆，合并进该馆 jsonl（不覆盖 by=user/ai 的结论）。"""
    with open(a.field_map, encoding='utf-8') as f:
        fmap = json.load(f)
    with open(a.json, encoding='utf-8') as f:
        data = json.load(f)
    records = extract_records(data, fmap.get('records_path'))
    fields = fmap.get('fields', {})

    catalog = load_catalog(a.catalog)
    if not a.no_freshness_check:
        check_freshness(a.catalog, len(catalog), dedup_dir)
    name_index, url_index = {}, {}
    for e in catalog:
        for cand in [e['name']] + e['aliases']:
            if cn := norm_name(cand):
                name_index.setdefault(cn, []).append(e)
        for u in e['urls']:
            url_index.setdefault(u, []).append(e)

    path = jsonl_path(dedup_dir, a.source_name)
    existing = read_jsonl(path)
    today = datetime.date.today().isoformat()
    out, tally, kept = [], {'已存在': 0, '疑似': 0, '未匹配': 0}, 0

    for i, d in enumerate(records):
        name = str(d.get(fields.get('name', 'name'), '') or '')
        raw_entry = str(d.get(fields.get('entry_url', ''), '') or '')
        brief = str(d.get(fields.get('description', ''), '') or '')
        urls = [u for k in ('entry_url', 'external_url')
                if (u := norm_url(d.get(fields.get(k, ''), '')))]
        rid = str(d.get(fields.get('id', ''), i + 1))
        tier, hits = match_one(name, urls, catalog, name_index, url_index, a.sim)
        tally[tier] += 1

        rec = existing.get(rid, {})
        if rec.get('by') in ('user', 'ai'):
            # 人和 AI 的结论不动，只刷新脚本部分
            rec.update({'name': name, 'entry': raw_entry, 'brief': brief, 'tier': tier,
                        'hits': hits[:5], 'source': a.source_name})
            kept += 1
        else:
            # 「已存在」和「未匹配」两档脚本能自己下结论；「疑似」留空等 AI 二审。
            # 未匹配给 70 分而不是 90：没匹配上不等于站上真没有，也可能是名称差太远。
            if tier == '已存在':
                verdict, wpid, conf = '已存在', (hits[0]['wpid'] if hits else None), 90
                reason = hits[0]['why'] if hits else ''
            elif tier == '未匹配':
                verdict, wpid, conf = '未收录', None, 70
                reason = '脚本未匹配到站内同名/同URL条目'
            else:
                verdict, wpid, conf, reason = '', None, None, ''
            rec = {
                'id': rid, 'source': a.source_name, 'name': name,
                'entry': raw_entry, 'brief': brief,
                'tier': tier, 'hits': hits[:5],
                'verdict': verdict, 'wpid': wpid, 'confidence': conf, 'reason': reason,
                'by': 'script', 'at': today, 'note': '',
            }
        out.append(rec)

    write_jsonl(path, out)
    print(f'总目 {len(catalog)} 条 × 采集 {len(records)} 条')
    print(f"已存在 {tally['已存在']} ｜ 疑似 {tally['疑似']} ｜ 未匹配 {tally['未匹配']}")
    if kept:
        print(f'保留了 {kept} 条已有的 AI/人工结论，未被脚本覆盖')
    print(f'结论 -> {path}')
    return tally


def main():
    p = argparse.ArgumentParser(description='入库查重：脚本粗筛 + 人工确认表（一条命令，跑几次都安全）')
    p.add_argument('--dedup-dir', required=True, help='查重目录，如 00-采集批次/02-.../02-查重')
    p.add_argument('--xlsx', help='批次总目 xlsx；给了就回收人工列并重新渲染「站内查重」sheet')
    p.add_argument('--json', help='该馆采集 JSON；给了就跑粗筛')
    p.add_argument('--field-map', help='配合 --json 的字段对应表')
    p.add_argument('--source-name', help='配合 --json 的馆名，决定 jsonl 文件名')
    p.add_argument('--catalog', default=DEFAULT_CATALOG, help='总目-与站点同步.xlsx')
    p.add_argument('--sim', type=float, default=0.75, help='名称相似度阈值（默认 0.75）')
    p.add_argument('--no-freshness-check', action='store_true',
                   help='跳过「快照是否过时」的联网检查')
    a = p.parse_args()

    if a.json and not (a.field_map and a.source_name):
        sys.exit('--json 需要同时给 --field-map 和 --source-name')
    if not a.json and not a.xlsx:
        sys.exit('至少要给 --json（跑粗筛）或 --xlsx（回收+渲染）之一')

    dedup_dir = a.dedup_dir
    os.makedirs(dedup_dir, exist_ok=True)

    # ① 先回收人工列，免得后面的写入把人的判断冲掉
    if a.xlsx and os.path.exists(a.xlsx):
        n = collect_manual(a.xlsx, dedup_dir)
        if n:
            print(f'已回收 {n} 条人工确认到 jsonl')

    # ② 粗筛
    if a.json:
        screen(a, dedup_dir)

    # ③ 重新渲染
    if a.xlsx:
        if not os.path.exists(a.xlsx):
            sys.exit(f'找不到 {a.xlsx}，先跑 build_batch_mulu.py 生成批次总目')
        n = render(a.xlsx, dedup_dir)
        print(f'「{SHEET}」sheet 已刷新，{n} 行 -> {a.xlsx}')
        print('人工确认：在「人工判定」列选（留空＝同意 AI 判定），改完再跑一次本命令即回收。')


if __name__ == '__main__':
    main()
